from osgeo import gdal
import openpyxl
import numpy as np

data_path =r'C:\Users\bangphuonglaptop\Desktop\python\dataImage.tif'

ds = gdal.Open(data_path, gdal.GA_ReadOnly)

transform = ds.GetGeoTransform() 
xOrigin = transform[0] 
yOrigin = transform[3] 
pixelWidth = transform[1] 
pixelHeight = transform[5] 

wb = openpyxl.load_workbook(r'C:\Users\bangphuonglaptop\Desktop\Sample.xlsx')
sheet = wb['Sheet1']

nband = 7
nsample = 120*9

for bandId in range(1,nband+1,1):
    band = ds.GetRasterBand(bandId)
    data = band.ReadAsArray()
    for row in range(2,nsample+2,1):
        x = sheet.cell(row=row, column=2).value
        y = sheet.cell(row=row, column=3).value
        xOffset = int((x - xOrigin) / pixelWidth)
        yOffset = int((y - yOrigin) / pixelHeight)
        value = data[yOffset][xOffset]
        sheet.cell(row=row, column=4+bandId, value=value)
        
for row in range(2,nsample+2,1):
    red = sheet.cell(row= row, column= 5).value
    green = sheet.cell(row= row, column= 6).value
    blue = sheet.cell(row= row, column= 7).value
    nir = sheet.cell(row= row, column= 8).value
    sheet.cell(row=row, column=12, value=(nir-red)/(nir+red))
    sheet.cell(row=row, column=13, value=(green-nir)/(green+nir))

wb.save(r'C:\Users\bangphuonglaptop\Desktop\Sample.xlsx')
